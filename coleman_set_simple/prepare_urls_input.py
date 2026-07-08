#!/usr/bin/env python3
"""
Normalize a user-provided URL list into the pipeline input CSV:
  brand, product_url

Supports input:
- .xlsx (multiple sheets; brand inferred from sheet name)
- .csv (any column name: product_url/url/link; brand optional)
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from typing import Dict, Iterable, List, Optional, Set, Tuple


# Match until whitespace or a common trailing delimiter.
URL_RE = re.compile(r"https?://[^\s)\"']+", re.IGNORECASE)


def _norm_brand(sheet_name: str) -> str:
    n = (sheet_name or "").strip().lower()
    if "cat" in n:
        return "catnapper"
    if "jack" in n:
        return "jackson"
    return re.sub(r"\\s+", "_", n) or "unknown"


def _looks_like_coleman_product(url: str) -> bool:
    u = (url or "").strip()
    if not u.lower().startswith("http"):
        return False
    if "colemanfurniture.com" not in u.lower():
        return False
    path = u.split("?", 1)[0].lower()
    return path.endswith((".htm", ".html"))


def _extract_urls_from_cell(val) -> List[str]:
    if val is None:
        return []
    if isinstance(val, str):
        text = val.strip()
    else:
        text = str(val).strip()
    if not text:
        return []
    urls = [m.group(0).strip() for m in URL_RE.finditer(text)]
    if not urls and text.lower().startswith("http"):
        urls = [text]
    return urls


def _from_xlsx(path: str) -> List[Tuple[str, str]]:
    try:
        import openpyxl
    except ModuleNotFoundError as e:
        raise RuntimeError("openpyxl missing. Install requirements.txt in workflow.") from e

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: List[Tuple[str, str]] = []
    seen: Set[str] = set()
    for sheet in wb.sheetnames:
        brand = _norm_brand(sheet)
        ws = wb[sheet]

        # Prefer header-based URL columns to avoid accidentally scanning other cells.
        header = []
        max_col = min(100, ws.max_column or 0)
        for c in range(1, max_col + 1):
            v = ws.cell(1, c).value
            header.append((c, (str(v).strip() if v is not None else "")))
        url_cols = [c for c, h in header if ("url" in h.lower() or "link" in h.lower())]
        if not url_cols:
            # Fallback: old behavior (scan all cells)
            url_cols = list(range(1, max_col + 1))

        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # skip header row
            if r_idx == 1:
                continue
            for c in url_cols:
                if c - 1 >= len(row):
                    continue
                cell = row[c - 1]
                for u in _extract_urls_from_cell(cell):
                    if not _looks_like_coleman_product(u):
                        continue
                    if u in seen:
                        continue
                    seen.add(u)
                    out.append((brand, u))
    return out


def _from_csv(path: str) -> List[Tuple[str, str]]:
    out: List[Tuple[str, str]] = []
    seen: Set[str] = set()
    with open(path, "r", encoding="utf-8") as f:
        r = csv.DictReader(f)
        fieldnames = [c.lower() for c in (r.fieldnames or [])]
        url_col = None
        for candidate in ("product_url", "url", "link"):
            if candidate in fieldnames:
                url_col = (r.fieldnames or [])[fieldnames.index(candidate)]
                break
        brand_col = None
        if "brand" in fieldnames:
            brand_col = (r.fieldnames or [])[fieldnames.index("brand")]

        if url_col is None:
            # fallback: take first column
            url_col = (r.fieldnames or ["url"])[0]

        for row in r:
            u = (row.get(url_col) or "").strip()
            b = (row.get(brand_col) or "").strip() if brand_col else ""
            if not u:
                continue
            if not _looks_like_coleman_product(u):
                continue
            if u in seen:
                continue
            seen.add(u)
            out.append((b or "unknown", u))
    return out


def write_out(rows: Iterable[Tuple[str, str]], out_path: str) -> None:
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["brand", "product_url"])
        w.writeheader()
        for brand, url in rows:
            w.writerow({"brand": brand, "product_url": url})


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="Input .xlsx or .csv")
    ap.add_argument("--out", required=True, help="Output CSV (brand, product_url)")
    args = ap.parse_args()

    inp = args.inp
    if not os.path.exists(inp):
        print(f"ERROR: input file not found: {inp}", file=sys.stderr)
        print(f"cwd: {os.getcwd()}", file=sys.stderr)
        try:
            print("repo root listing:", file=sys.stderr)
            for name in sorted(os.listdir("."))[:200]:
                print(" -", name, file=sys.stderr)
        except Exception:
            pass
        return 2
    low = inp.lower()
    if low.endswith(".xlsx") or low.endswith(".xlsm"):
        rows = _from_xlsx(inp)
    elif low.endswith(".csv"):
        rows = _from_csv(inp)
    else:
        print("Unsupported input type. Use .xlsx or .csv", file=sys.stderr)
        return 2

    write_out(rows, args.out)
    print(f"Saved {len(rows)} urls -> {args.out}")
    return 0 if rows else 1


if __name__ == "__main__":
    raise SystemExit(main())

